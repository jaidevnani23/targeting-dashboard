#!/usr/bin/env python3
"""
Product Discovery Script
========================
Scrapes trending/bestselling products from Amazon India and Flipkart
via BrightData Web Access API (serp_trends1), compares against existing
products in Demand_Excel_Filled.xlsx, reads supplier xlsx files to
determine which states have meaningful supplier presence for each
product's NIC code, and outputs new_products_suggestions.xlsx.

BrightData Web Access API:
  - Find your API key: BrightData dashboard > Web Access > serp_trends1
  - Paste it into BD_WEB_ACCESS_KEY below
  - Zone name is: serp_trends1
  - Cost: $1.50/CPM (~$0.0015 per request)

After running:
  1. Open new_products_suggestions.xlsx
  2. Fill in "Search Term" for products you want to keep
  3. Delete rows you don't want
  4. Upload the file via the dashboard (Product Review tab)
     OR run product_updater.py directly

Requirements:
    pip install requests pandas openpyxl beautifulsoup4 lxml python-dotenv
"""

import requests
import pandas as pd
import json
import os
import time
import logging
import random
from bs4 import BeautifulSoup
from datetime import datetime
from collections import defaultdict

# ── .env support ──────────────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ── CONFIG — fill these in ────────────────────────────────────────────────────
DEMAND_FILE    = "data/Demand_Excel_Filled.xlsx"
NIC_CODES_FILE = "data/Key_NIC_Codes_List.xlsx"
SUPPLIERS_DIR  = "data/suppliers"
OUTPUT_FILE    = "data/new_products_suggestions.xlsx"

# BrightData Web Access API key
# Find in: BrightData dashboard > Web Access > serp_trends1 > API Key
BD_WEB_ACCESS_KEY  = os.environ.get("BD_WEB_ACCESS_KEY", "7e2f85a0-e82b-420d-b5f1-fe10f1fe9774")
BD_WEB_ACCESS_ZONE = "serp_trends1"
BD_WEB_ACCESS_URL  = "https://api.brightdata.com/request"

SKIP_AMAZON   = False
SKIP_FLIPKART = False

MAX_PER_CATEGORY = 10
MIN_FLOOR_STATES = 3

MIN_SCRAPE_DELAY   = 2.0
MAX_SCRAPE_DELAY   = 4.0
MIN_CATEGORY_DELAY = 6.0
MAX_CATEGORY_DELAY = 10.0

MAX_RETRIES   = 3
RETRY_BACKOFF = 10

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s  %(levelname)-7s  %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

# ── BRIGHTDATA WEB ACCESS API ─────────────────────────────────────────────────
def _credentials_ok() -> bool:
    return BD_WEB_ACCESS_KEY not in {"YOUR_API_KEY_HERE", ""}

def _fetch_url(url: str, site_name: str) -> str | None:
    headers = {
        "Content-Type":  "application/json",
        "Authorization": f"Bearer {BD_WEB_ACCESS_KEY}",
    }
    payload = {
        "zone":   BD_WEB_ACCESS_ZONE,
        "url":    url,
        "format": "raw",
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.post(
                BD_WEB_ACCESS_URL,
                headers=headers,
                json=payload,
                timeout=30,
            )

            if resp.status_code == 200:
                return resp.text
            elif resp.status_code == 401:
                log.error(
                    "Web Access API returned 401 — API key is invalid or zone is disabled.\n"
                    "  -> Check BrightData > Web Access > serp_trends1 is Active\n"
                    "  -> Re-copy the API key into BD_WEB_ACCESS_KEY"
                )
                return None
            elif resp.status_code == 429:
                log.warning(f"  [{site_name}] Rate limited (attempt {attempt}/{MAX_RETRIES})")
            elif resp.status_code in (503, 502):
                log.warning(f"  [{site_name}] HTTP {resp.status_code} (attempt {attempt}/{MAX_RETRIES})")
            else:
                log.warning(f"  [{site_name}] HTTP {resp.status_code} — skipping")
                return None

        except Exception as e:
            log.warning(f"  [{site_name}] Request error (attempt {attempt}/{MAX_RETRIES}): {e}")

        if attempt < MAX_RETRIES:
            log.info(f"  Backing off {RETRY_BACKOFF}s before retry...")
            time.sleep(RETRY_BACKOFF)

    log.warning(f"  [{site_name}] All {MAX_RETRIES} attempts failed — skipping keyword")
    return None


def scrape_amazon(search_term: str) -> list:
    if SKIP_AMAZON:
        return []
    url  = f"https://www.amazon.in/s?k={requests.utils.quote(search_term)}&s=review-rank"
    html = _fetch_url(url, "Amazon")
    if not html:
        return []
    soup     = BeautifulSoup(html, "lxml")
    products = []
    for tag in soup.select("span.a-text-normal"):
        text = tag.get_text(strip=True)
        if 10 < len(text) < 120:
            products.append(text)
    log.info(f"  Amazon: {len(products[:15])} products for '{search_term}'")
    return products[:15]


def scrape_flipkart(search_term: str) -> list:
    if SKIP_FLIPKART:
        return []
    url  = f"https://www.flipkart.com/search?q={requests.utils.quote(search_term)}&sort=popularity"
    html = _fetch_url(url, "Flipkart")
    if not html:
        return []
    soup     = BeautifulSoup(html, "lxml")
    products = []
    for selector in ["div._4rR01T", "a.s1Q9rs", "div.KzDlHZ", "div.col-7-12 a"]:
        for tag in soup.select(selector):
            text = tag.get_text(strip=True)
            if 10 < len(text) < 120:
                products.append(text)
    log.info(f"  Flipkart: {len(products[:15])} products for '{search_term}'")
    return products[:15]


def random_scrape_delay():
    delay = random.uniform(MIN_SCRAPE_DELAY, MAX_SCRAPE_DELAY)
    log.info(f"  Waiting {delay:.1f}s...")
    time.sleep(delay)

def random_category_delay():
    delay = random.uniform(MIN_CATEGORY_DELAY, MAX_CATEGORY_DELAY)
    log.info(f"Category done. Waiting {delay:.1f}s before next...")
    time.sleep(delay)

# ── LOAD REFERENCE DATA ───────────────────────────────────────────────────────
def load_existing_products() -> set:
    df       = pd.read_excel(DEMAND_FILE)
    prod_col = next(c for c in df.columns if 'product' in c.lower())
    return set(df[prod_col].str.lower().str.strip().tolist())


def load_nic_reference() -> tuple:
    nic_df    = pd.read_excel(NIC_CODES_FILE)
    demand_df = pd.read_excel(DEMAND_FILE)

    code_col  = next(c for c in nic_df.columns if 'nic' in c.lower() and 'code' in c.lower())
    desc_col  = next(c for c in nic_df.columns if 'desc' in c.lower())
    nic_col_d = next(c for c in demand_df.columns if 'nic' in c.lower())
    cat_col_d = next(c for c in demand_df.columns if 'cat' in c.lower())

    nic_df[code_col]     = nic_df[code_col].astype(str).str.strip()
    demand_df[nic_col_d] = demand_df[nic_col_d].astype(str).str.strip()

    nic_lookup = dict(zip(nic_df[code_col], nic_df[desc_col]))
    cat_lookup = (
        demand_df.groupby(nic_col_d)[cat_col_d]
        .agg(lambda x: x.mode()[0])
        .to_dict()
    )
    log.info(f"Loaded {len(nic_lookup)} NIC codes, {len(cat_lookup)} category mappings")
    return nic_lookup, cat_lookup


def load_categories_and_keywords(nic_lookup: dict, cat_lookup: dict) -> dict:
    demand_df  = pd.read_excel(DEMAND_FILE)
    nic_col_d  = next(c for c in demand_df.columns if 'nic' in c.lower())
    cat_col_d  = next(c for c in demand_df.columns if 'cat' in c.lower())
    prod_col_d = next(c for c in demand_df.columns if 'product' in c.lower())
    demand_df[nic_col_d] = demand_df[nic_col_d].astype(str).str.strip()
    nic_to_cat = dict(zip(demand_df[nic_col_d], demand_df[cat_col_d]))

    cat_data = {}
    for nic_code, nic_desc in nic_lookup.items():
        category = nic_to_cat.get(nic_code, nic_desc)
        keywords = _keywords_from_description(nic_desc)
        if category not in cat_data:
            cat_data[category] = {"nic_codes": [], "keywords": set()}
        cat_data[category]["nic_codes"].append(nic_code)
        cat_data[category]["keywords"].update(keywords)

    for _, row in demand_df.iterrows():
        cat     = row[cat_col_d]
        product = str(row[prod_col_d]).strip()
        if cat in cat_data:
            cat_data[cat]["keywords"].add(product.lower())

    return cat_data


def _keywords_from_description(desc: str) -> list:
    stop = {"of", "and", "or", "the", "in", "via", "other", "n.e.c",
            "not", "stores", "stalls", "markets", "retail", "sale",
            "wholesale", "manufacture", "articles", "related"}
    words    = desc.lower().replace(",", " ").replace(".", " ").split()
    keywords = [w for w in words if w not in stop and len(w) > 3]
    keywords.append(desc.lower())
    return keywords

# ── SUPPLIER STATE ANALYSIS ───────────────────────────────────────────────────
def load_supplier_counts_by_nic() -> dict:
    nic_state_counts = defaultdict(lambda: defaultdict(int))

    if not os.path.exists(SUPPLIERS_DIR):
        log.warning(f"Suppliers directory not found: {SUPPLIERS_DIR} — state allocation skipped")
        return {}

    files = [f for f in os.listdir(SUPPLIERS_DIR) if f.endswith('.xlsx')]
    if not files:
        log.warning("No supplier xlsx files found — state allocation skipped")
        return {}

    log.info(f"Reading {len(files)} supplier state files...")
    for fname in files:
        path = os.path.join(SUPPLIERS_DIR, fname)
        try:
            df = pd.read_excel(path, dtype=str)
            for _, record in df.iterrows():
                nic_code = str(record.get("NIC_Code", "")).strip()
                state    = str(record.get("State",    "")).strip().title()
                if nic_code and state and nic_code != 'nan' and state != 'nan':
                    nic_state_counts[nic_code][state] += 1
        except Exception as e:
            log.warning(f"Could not read {fname}: {e}")

    log.info(f"Loaded supplier counts for {len(nic_state_counts)} NIC codes")
    return dict(nic_state_counts)


def get_states_for_product(nic_code: str, nic_state_counts: dict) -> list:
    state_counts = nic_state_counts.get(nic_code, {})
    if not state_counts:
        return []

    sorted_states = sorted(state_counts.items(), key=lambda x: x[1], reverse=True)
    total_states  = len(sorted_states)
    tier1_cutoff  = max(1, int(total_states * 0.20))
    tier2_cutoff  = max(2, int(total_states * 0.50))

    selected = []
    for i, (state, count) in enumerate(sorted_states):
        tier = 1 if i < tier1_cutoff else (2 if i < tier2_cutoff else 3)
        if tier <= 2:
            selected.append({"state": state, "supplier_count": count, "tier": tier})

    if len(selected) < MIN_FLOOR_STATES:
        already = {s["state"] for s in selected}
        for state, count in sorted_states:
            if len(selected) >= MIN_FLOOR_STATES:
                break
            if state not in already:
                selected.append({"state": state, "supplier_count": count, "tier": 3})

    return selected

# ── DISCOVERY ─────────────────────────────────────────────────────────────────
def discover_new_products(cat_data: dict, existing: set,
                          nic_state_counts: dict, cat_lookup: dict) -> list:
    suggestions      = []
    total_categories = len(cat_data)
    category_num     = 0

    for category, data in cat_data.items():
        category_num += 1
        nic_codes = data["nic_codes"]
        keywords  = list(data["keywords"])[:3]
        log.info(f"\n[{category_num}/{total_categories}] Searching: {category} ({len(nic_codes)} NIC codes)")
        found = []

        for keyword in keywords:
            log.info(f"  Keyword: '{keyword}'")

            amazon_products   = scrape_amazon(keyword)
            random_scrape_delay()

            flipkart_products = scrape_flipkart(keyword)
            random_scrape_delay()

            for product in amazon_products + flipkart_products:
                p_lower = product.lower().strip()
                if any(e in p_lower or p_lower in e for e in existing):
                    continue
                if any(s["Product"].lower() == p_lower for s in found):
                    continue

                nic_code         = nic_codes[0]
                state_allocation = get_states_for_product(nic_code, nic_state_counts)

                found.append({
                    "Product":              product,
                    "Category":             category,
                    "NIC_Code":             nic_code,
                    "Search Term":          "",          # ← user fills this in
                    "Source":               "Amazon" if product in amazon_products else "Flipkart",
                    "Keyword Used":         keyword,
                    "States":               " | ".join(s["state"] for s in state_allocation),
                    "States_Count":         len(state_allocation),
                    # Keep full allocation as JSON string for product_updater.py
                    "State_Allocation_JSON": json.dumps(state_allocation),
                })

        suggestions.extend(found[:MAX_PER_CATEGORY])
        log.info(f"  Found {len(found)} new products for {category}")

        if category_num < total_categories:
            random_category_delay()

    return suggestions

# ── WRITE XLSX ────────────────────────────────────────────────────────────────
def write_xlsx(suggestions: list):
    """
    Writes suggestions to an XLSX file.
    Columns visible to the reviewer:
        Product | Category | NIC_Code | Search Term | Source |
        Keyword Used | States | States_Count
    State_Allocation_JSON is written to a hidden-ish last column so
    product_updater.py can reconstruct the full state list without
    needing to re-run discovery.
    """
    if not suggestions:
        log.warning("No suggestions to write.")
        return

    # Column order — Search Term is col D so it's easy to fill
    col_order = [
        "Product",
        "Category",
        "NIC_Code",
        "Search Term",
        "Source",
        "Keyword Used",
        "States",
        "States_Count",
        "State_Allocation_JSON",   # kept for updater; reviewer can ignore
    ]

    df = pd.DataFrame(suggestions)[col_order]

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Product Suggestions")

        ws = writer.sheets["Product Suggestions"]

        # ── Column widths ──────────────────────────────────────────────────
        widths = {
            "A": 50,  # Product
            "B": 30,  # Category
            "C": 12,  # NIC_Code
            "D": 35,  # Search Term  ← highlighted
            "E": 12,  # Source
            "F": 30,  # Keyword Used
            "G": 60,  # States
            "H": 14,  # States_Count
            "I": 20,  # State_Allocation_JSON (narrow — machine use only)
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # ── Highlight "Search Term" header and cells ───────────────────────
        from openpyxl.styles import PatternFill, Font, Alignment
        highlight_fill   = PatternFill("solid", fgColor="FFF3CD")   # amber tint
        header_fill      = PatternFill("solid", fgColor="0D2240")   # navy
        header_font      = Font(color="FFFFFF", bold=True)
        search_hdr_fill  = PatternFill("solid", fgColor="1A56B8")   # blue
        search_cell_fill = PatternFill("solid", fgColor="EBF2FF")   # light blue

        # Style all header cells
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Search Term header gets its own colour
        ws["D1"].fill = search_hdr_fill

        # Highlight all Search Term cells so reviewer knows where to type
        for row_idx in range(2, len(suggestions) + 2):
            ws[f"D{row_idx}"].fill      = search_cell_fill
            ws[f"D{row_idx}"].alignment = Alignment(horizontal="left")

        # Freeze top row
        ws.freeze_panes = "A2"

    log.info(f"Saved {len(suggestions)} suggestions → {OUTPUT_FILE}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("Product Discovery — Quarterly Run")
    log.info(f"Date      : {datetime.now().strftime('%Y-%m-%d')}")
    log.info(f"Output    : XLSX (new_products_suggestions.xlsx)")
    log.info(f"Sources   : {'Flipkart only' if SKIP_AMAZON else 'Amazon India + Flipkart'}"
             + " via BrightData Web Access API")
    log.info("=" * 60)

    if not _credentials_ok():
        log.error(
            "BD_WEB_ACCESS_KEY not set.\n"
            "  1. Go to BrightData dashboard > Web Access > serp_trends1\n"
            "  2. Make sure the zone is Active (toggle top right)\n"
            "  3. Copy the API Key and paste into BD_WEB_ACCESS_KEY above"
        )
        return

    existing               = load_existing_products()
    nic_lookup, cat_lookup = load_nic_reference()
    cat_data               = load_categories_and_keywords(nic_lookup, cat_lookup)
    nic_state_counts       = load_supplier_counts_by_nic()

    log.info(f"Existing products           : {len(existing)}")
    log.info(f"Categories                  : {len(cat_data)}")
    log.info(f"NIC codes with supplier data: {len(nic_state_counts)}"
             + (" (run msme_supplier_fetcher.py to populate)" if not nic_state_counts else ""))

    suggestions = discover_new_products(cat_data, existing, nic_state_counts, cat_lookup)

    write_xlsx(suggestions)

    total_rows = sum(
        len(json.loads(s["State_Allocation_JSON"])) for s in suggestions
        if s.get("State_Allocation_JSON")
    )

    log.info("\n" + "=" * 60)
    log.info(f"Total suggestions          : {len(suggestions)}")
    log.info(f"Total rows if all approved : {total_rows}")
    log.info(f"Avg states per product     : {total_rows // max(len(suggestions), 1)}")
    log.info(f"Saved to                   : {OUTPUT_FILE}")
    log.info("")
    log.info("Next steps:")
    log.info("  1. Open new_products_suggestions.xlsx")
    log.info("  2. Fill in column D 'Search Term' for products you want")
    log.info("  3. Delete any rows you don't want")
    log.info("  4. Upload the file via the dashboard (Product Review tab)")
    log.info("     OR run product_updater.py directly")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
